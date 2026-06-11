from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.brands.models import Brand
from app.common.utils import calculate_price_difference, normalize_part_number
from app.config import settings
from app.files.models import UploadedFile, UploadedFileStatus
from app.files.schemas import UploadedFileProgressUpdate
from app.files.service import (
    get_processed_file_path,
    set_processed_file_path,
    update_file_progress,
    update_file_status,
)
from app.logs.schemas import ProcessingLogCreateData
from app.logs.service import create_processing_log
from app.lookup.service import PartLookupService
from app.parts.models import Part
from app.parts.service import (
    create_part_from_lookup,
    get_part_by_brand_and_number,
    update_part_price,
)
from app.prices.schemas import PriceHistoryCreateData
from app.prices.service import create_price_history
from app.processing.utils import (
    clean_part_number,
    detect_part_number_column,
    read_uploaded_file,
)


class ExcelProcessor:
    def __init__(self, db: Session):
        self.db = db
        self.lookup_service = PartLookupService()

    def process_uploaded_file(self, uploaded_file: UploadedFile) -> UploadedFile:
        try:
            update_file_status(
                db=self.db,
                uploaded_file=uploaded_file,
                status=UploadedFileStatus.PROCESSING,
            )

            brand = uploaded_file.brand

            if not brand:
                raise ValueError("Brand not found for uploaded file")

            df = read_uploaded_file(uploaded_file.original_file_path)

            if df.empty:
                raise ValueError("Uploaded file is empty")

            part_number_column = detect_part_number_column(list(df.columns))

            if not part_number_column:
                raise ValueError("Part number column not found")

            extracted_rows = self._extract_part_rows(
                df=df,
                part_number_column=part_number_column,
            )

            total_rows = len(extracted_rows)

            counters = {
                "processed_rows": 0,
                "found_count": 0,
                "not_found_count": 0,
                "price_changed_count": 0,
            }

            update_file_progress(
                db=self.db,
                uploaded_file=uploaded_file,
                progress=UploadedFileProgressUpdate(
                    total_rows=total_rows,
                    processed_rows=0,
                    found_count=0,
                    not_found_count=0,
                    price_changed_count=0,
                ),
            )

            existing_part_map, new_part_numbers = self._split_existing_and_new_parts(
                brand=brand,
                extracted_rows=extracted_rows,
            )

            new_part_lookup_results = self.lookup_service.lookup_new_parts_batch(
                brand_name=brand.name,
                part_numbers=new_part_numbers,
            )

            existing_price_lookup_results = self.lookup_service.lookup_market_prices_batch(
                brand_name=brand.name,
                parts=[
                    {
                        "part_number": part.part_number,
                        "product_name": part.product_name,
                    }
                    for part in existing_part_map.values()
                ],
            )

            processed_part_map: dict[str, Part] = {}

            for part_number, part in existing_part_map.items():
                processed_part_map[part_number] = self._process_existing_part(
                    uploaded_file=uploaded_file,
                    brand=brand,
                    part=part,
                    price_lookup=existing_price_lookup_results.get(part_number, {}),
                    counters=counters,
                )

            for part_number in new_part_numbers:
                processed_part_map[part_number] = self._process_new_part(
                    uploaded_file=uploaded_file,
                    brand=brand,
                    part_number=part_number,
                    lookup_data=new_part_lookup_results.get(part_number, {}),
                    counters=counters,
                )

            output_rows = []

            for item in extracted_rows:
                part_number = item["part_number"]
                row_number = item["row_number"]
                part = processed_part_map.get(part_number)

                if not part:
                    counters["processed_rows"] += 1
                    counters["not_found_count"] += 1

                    self._create_log(
                        uploaded_file=uploaded_file,
                        brand_id=brand.id,
                        part_number=part_number,
                        row_number=row_number,
                        status="failed",
                        message="Part was not processed",
                    )

                    output_rows.append(
                        self._build_output_row(
                            brand_name=brand.name,
                            part_number=part_number,
                            product_name="Not Found",
                            description="Part was not processed",
                            current_price=None,
                            previous_price=None,
                            currency=settings.DEFAULT_CURRENCY,
                            price_changed=False,
                            source_url=None,
                            lookup_status="failed",
                        )
                    )

                    continue

                output_rows.append(
                    self._build_output_row_from_part(
                        brand_name=brand.name,
                        part=part,
                    )
                )

                counters["processed_rows"] += 1

                update_file_progress(
                    db=self.db,
                    uploaded_file=uploaded_file,
                    progress=UploadedFileProgressUpdate(
                        total_rows=total_rows,
                        processed_rows=counters["processed_rows"],
                        found_count=counters["found_count"],
                        not_found_count=counters["not_found_count"],
                        price_changed_count=counters["price_changed_count"],
                    ),
                )

            processed_file_path = self._generate_processed_excel(
                uploaded_file=uploaded_file,
                output_rows=output_rows,
            )

            set_processed_file_path(
                db=self.db,
                uploaded_file=uploaded_file,
                processed_file_path=processed_file_path,
            )

            update_file_status(
                db=self.db,
                uploaded_file=uploaded_file,
                status=UploadedFileStatus.COMPLETED,
            )

            self.db.refresh(uploaded_file)
            return uploaded_file

        except Exception as exc:
            update_file_status(
                db=self.db,
                uploaded_file=uploaded_file,
                status=UploadedFileStatus.FAILED,
                error_message=str(exc),
            )

            self._create_log(
                uploaded_file=uploaded_file,
                brand_id=uploaded_file.brand_id,
                part_number=None,
                row_number=None,
                status="failed",
                message=f"File processing failed: {str(exc)}",
            )

            raise

    def _extract_part_rows(
        self,
        df: pd.DataFrame,
        part_number_column: str,
    ) -> list[dict[str, Any]]:
        extracted_rows = []

        for index, row in df.iterrows():
            part_number = clean_part_number(row.get(part_number_column))

            if not part_number:
                continue

            extracted_rows.append(
                {
                    "row_number": index + 2,
                    "part_number": normalize_part_number(part_number),
                }
            )

        return extracted_rows

    def _split_existing_and_new_parts(
        self,
        brand: Brand,
        extracted_rows: list[dict[str, Any]],
    ) -> tuple[dict[str, Part], list[str]]:
        unique_part_numbers = []

        for item in extracted_rows:
            part_number = item["part_number"]

            if part_number not in unique_part_numbers:
                unique_part_numbers.append(part_number)

        existing_part_map: dict[str, Part] = {}
        new_part_numbers: list[str] = []

        for part_number in unique_part_numbers:
            existing_part = get_part_by_brand_and_number(
                db=self.db,
                brand_id=brand.id,
                part_number=part_number,
            )

            if existing_part:
                existing_part_map[part_number] = existing_part
            else:
                new_part_numbers.append(part_number)

        return existing_part_map, new_part_numbers

    def _process_existing_part(
        self,
        uploaded_file: UploadedFile,
        brand: Brand,
        part: Part,
        price_lookup: dict[str, Any],
        counters: dict[str, int],
    ) -> Part:
        old_price = part.current_price

        new_price = price_lookup.get("market_price")
        currency = price_lookup.get("currency")
        source_url = price_lookup.get("source_url")
        lookup_status = price_lookup.get("lookup_status") or "price_not_available"

        updated_part, price_changed = update_part_price(
            db=self.db,
            part=part,
            new_price=new_price,
            currency=currency,
            source_url=source_url,
            lookup_status=lookup_status,
        )

        create_price_history(
            db=self.db,
            data=PriceHistoryCreateData(
                part_id=updated_part.id,
                uploaded_file_id=uploaded_file.id,
                old_price=old_price,
                new_price=new_price,
                currency=currency or updated_part.currency,
                source_url=source_url,
                price_changed=price_changed,
            ),
        )

        if price_changed:
            counters["price_changed_count"] += 1

        counters["found_count"] += 1

        self._create_log(
            uploaded_file=uploaded_file,
            brand_id=brand.id,
            part_number=part.part_number,
            row_number=None,
            status="price_checked",
            message="Existing part price checked using batch lookup",
        )

        return updated_part

    def _process_new_part(
        self,
        uploaded_file: UploadedFile,
        brand: Brand,
        part_number: str,
        lookup_data: dict[str, Any],
        counters: dict[str, int],
    ) -> Part:
        new_part = create_part_from_lookup(
            db=self.db,
            brand_id=brand.id,
            part_number=part_number,
            lookup_data=lookup_data,
        )

        create_price_history(
            db=self.db,
            data=PriceHistoryCreateData(
                part_id=new_part.id,
                uploaded_file_id=uploaded_file.id,
                old_price=None,
                new_price=new_part.current_price,
                currency=new_part.currency,
                source_url=new_part.source_url,
                price_changed=False,
            ),
        )

        if new_part.lookup_status == "found":
            counters["found_count"] += 1
            log_status = "found"
            message = "New part added using batch lookup"
        else:
            counters["not_found_count"] += 1
            log_status = "not_found"
            message = "New part not found using batch lookup"

        self._create_log(
            uploaded_file=uploaded_file,
            brand_id=brand.id,
            part_number=part_number,
            row_number=None,
            status=log_status,
            message=message,
        )

        return new_part

    def _build_output_row_from_part(
        self,
        brand_name: str,
        part: Part,
    ) -> dict[str, Any]:
        price_changed = False

        if part.previous_price is not None and part.current_price is not None:
            price_changed = Decimal(str(part.previous_price)) != Decimal(str(part.current_price))

        return self._build_output_row(
            brand_name=brand_name,
            part_number=part.part_number,
            product_name=part.product_name,
            description=part.description,
            current_price=part.current_price,
            previous_price=part.previous_price,
            currency=part.currency,
            price_changed=price_changed,
            source_url=part.source_url,
            lookup_status=part.lookup_status,
        )

    def _build_output_row(
        self,
        brand_name: str,
        part_number: str,
        product_name: str | None,
        description: str | None,
        current_price: Decimal | None,
        previous_price: Decimal | None,
        currency: str | None,
        price_changed: bool,
        source_url: str | None,
        lookup_status: str,
    ) -> dict[str, Any]:
        return {
            "Brand": brand_name,
            "Part Number": part_number,
            "Product Name": product_name or "Not Found",
            "Description": description or "Not Found",
            "Current Price": current_price,
            "Previous Price": previous_price,
            "Price Difference": calculate_price_difference(
                current_price=current_price,
                previous_price=previous_price,
            ),
            "Currency": currency or settings.DEFAULT_CURRENCY,
            "Price Changed": "Yes" if price_changed else "No",
            "Source URL": source_url or "",
            "Lookup Status": lookup_status,
        }

    def _generate_processed_excel(
        self,
        uploaded_file: UploadedFile,
        output_rows: list[dict[str, Any]],
    ) -> str:
        processed_filename = (
            f"processed_{uploaded_file.id}_{Path(uploaded_file.original_filename).stem}.xlsx"
        )

        processed_file_path = get_processed_file_path(processed_filename)

        Path(settings.PROCESSED_FILE_DIR).mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(processed_file_path, engine="openpyxl") as writer:
            pd.DataFrame(output_rows).to_excel(
                writer,
                sheet_name="Processed Parts",
                index=False,
            )

            metadata = pd.DataFrame(
                [
                    {
                        "Generated At": datetime.now(timezone.utc).isoformat(),
                        "Uploaded File ID": uploaded_file.id,
                        "Original Filename": uploaded_file.original_filename,
                        "Status": "completed",
                        "Lookup Mode": "batch",
                        "Batch Size": settings.LOOKUP_BATCH_SIZE,
                    }
                ]
            )

            metadata.to_excel(
                writer,
                sheet_name="Metadata",
                index=False,
            )

        self._format_processed_excel(processed_file_path)

        return processed_file_path

    def _format_processed_excel(self, file_path: str) -> None:
        workbook = load_workbook(file_path)

        worksheet = workbook["Processed Parts"]

        header_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid",
        )

        not_found_fill = PatternFill(
            start_color="F8D7DA",
            end_color="F8D7DA",
            fill_type="solid",
        )

        price_changed_fill = PatternFill(
            start_color="FFF3CD",
            end_color="FFF3CD",
            fill_type="solid",
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_map = {cell.value: cell.column for cell in worksheet[1]}

        lookup_status_col = header_map.get("Lookup Status")
        price_changed_col = header_map.get("Price Changed")

        for row in worksheet.iter_rows(min_row=2):
            row_index = row[0].row

            lookup_status_value = (
                worksheet.cell(row=row_index, column=lookup_status_col).value
                if lookup_status_col
                else None
            )

            price_changed_value = (
                worksheet.cell(row=row_index, column=price_changed_col).value
                if price_changed_col
                else None
            )

            if str(lookup_status_value).lower() in ["not_found", "failed"]:
                for cell in row:
                    cell.fill = not_found_fill

            if str(price_changed_value).lower() == "yes":
                for cell in row:
                    cell.fill = price_changed_fill

        for column_name in ["Current Price", "Previous Price", "Price Difference"]:
            column_index = header_map.get(column_name)

            if column_index:
                for col in worksheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                ):
                    for item in col:
                        if isinstance(item.value, (int, float, Decimal)):
                            item.number_format = '#,##0.00'

        for ws in workbook.worksheets:
            for column_cells in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )

                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                    max_length + 2,
                    60,
                )

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        workbook.save(file_path)

    def _create_log(
        self,
        uploaded_file: UploadedFile,
        brand_id: int,
        part_number: str | None,
        row_number: int | None,
        status: str,
        message: str,
    ) -> None:
        create_processing_log(
            db=self.db,
            data=ProcessingLogCreateData(
                uploaded_file_id=uploaded_file.id,
                brand_id=brand_id,
                part_number=part_number,
                row_number=row_number,
                status=status,
                message=message,
            ),
        )